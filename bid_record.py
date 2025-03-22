from datetime import datetime
import os
from urllib.parse import urlparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.hyperlink import Hyperlink
from utils import compare_urls

class BidRecord:
    def __init__(self):
        self.bid_record_h = [
            "No", "Site", "Title", "Company", "Job Detail", "Company Url", "Start", "End", 
            "Bid Duration", "Resume"
        ]
        self.CheckAndCreateExcel()

    # Modified Exist function with try-except block
    def Exist(self, title: str, company: str, url: str, company_url: str) -> int:
        try:
            if not os.path.exists("bid_record.xlsx"):  # Check if file exists before attempting to load
                raise FileNotFoundError(f"The file bid_record.xlsx does not exist.")
            
            wb = load_workbook("bid_record.xlsx")
            sheet = wb.active

            # Parse the company_url to get the main DNS address
            parsed_url = urlparse(company_url)
            main_dns = f"{parsed_url.scheme}://{parsed_url.netloc}"

            # Iterate over the rows (skipping the header)
            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=len(self.bid_record_h)):
                job_detail = sheet.cell(row=row[4].row, column=5).value  # Assuming "Job Detail" is column 5
                company_name = sheet.cell(row=row[4].row, column=4).value  # Assuming "Company" is column 4
                company_url_value = sheet.cell(row=row[4].row, column=6).value  # Assuming "Company Url" is column 6
                job_title = sheet.cell(row=row[4].row, column=3).value  # Assuming "Title" is column 3

                # If the "Job Detail" column contains the URL
                if job_detail and compare_urls(url, job_detail):
                    return { 'code': 1, 'url': job_detail, 'title': job_title, 'company_name': company_name, 'company_url': company_url_value, 'no': row[0].row } 

                # If the "Company" column matches the company name OR the company_url matches the main DNS
                if company_name and company_name == company:
                    return { 'code': 2, 'url': job_detail, 'title': job_title, 'company_name': company_name, 'company_url': company_url_value, 'no': row[0].row } 
                                                                                                                                                                                                
                # If the "Title" matches the job title
                if job_title and job_title == title:
                    return { 'code': 4, 'url': job_detail, 'title': job_title, 'company_name': company_name, 'company_url': company_url_value, 'no': row[0].row } 

            return { 'code': 0 } 
        
        except FileNotFoundError as fnf_error:
            print(f"File not found error: {fnf_error}")
            return { 'code': -2 } 

        except Exception as e:
            print(f"An error occurred in the Exist method: {e}")
            return { 'code': -1 } 
    
    # Function to check and create Excel file (UpperCamelCase)
    def CheckAndCreateExcel(self):
        if os.path.exists("bid_record.xlsx"):
            wb = load_workbook("bid_record.xlsx")
            sheet = wb.active
            if sheet['A1'].value != "uranus":
                os.remove("bid_record.xlsx")
                self.CreateExcel()
        else:
            self.CreateExcel()

    # Function to create Excel file (UpperCamelCase)
    def CreateExcel(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Sheet1'

        for col_num, header in enumerate(self.bid_record_h, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)

        sheet['A1'] = 'uranus'
        wb.save("bid_record.xlsx")

    def AddRecord(self, job_data: dict):
        wb = load_workbook("bid_record.xlsx")
        sheet = wb.active

        next_row = sheet.max_row + 1  # Find next available row

        # Set 'No' column (autoincrement value)
        sheet.cell(row=next_row, column=1).value = next_row - 1  # Autoincrement the 'No' value

        # Insert the job_data values sequentially into the row (skip 'No' column)
        for col_num, key in enumerate(self.bid_record_h[1:], 2):  # Start from the second column, skipping 'No'
            if key in job_data:
                cell = sheet.cell(row=next_row, column=col_num)

                cell.value = job_data[key]
                                
                # Check if the column is a hyperlink and apply the appropriate styling
                if key in ['Site', 'Job Detail', 'Company Url', 'Resume']:
                    # Apply blue, underlined style for hyperlinks
                    cell.font = Font(color="0000FF", underline="single")
                    # Add hyperlink functionality for the respective columns
                    if key == 'Site':
                        cell.hyperlink = job_data[key]  # Assuming job_data[key] is the URL for Site
                    elif key == 'Job Detail':
                        cell.hyperlink = job_data[key]  # Assuming job_data[key] is the Job Detail URL
                    elif key == 'Company Url':
                        cell.hyperlink = job_data[key]  # Assuming job_data[key] is the Company URL
                    elif key == 'Resume':
                        cell.hyperlink = f'file:///{job_data[key]}'  # Assuming job_data[key] is the local file path for Resume

        wb.save("bid_record.xlsx")

        # Return the row number ('No') of the added record
        return next_row - 1

    # Function to modify an existing record based on 'No' (row number) (UpperCamelCase)
    def FinalizeRecord(self, no: int):
        wb = load_workbook("bid_record.xlsx")
        sheet = wb.active

        row = no + 1  # Adjust for 1-based row indexing in Excel
        if row > sheet.max_row:
            raise ValueError(f"Record with No {no} does not exist.")

        # Update the values in the row corresponding to the 'No'
        for col_num, key in enumerate(self.bid_record_h[1:], 2):  # Start from second column, skipping 'No'
            cell = sheet.cell(row=row, column=col_num)
            
            # Check if the column is a hyperlink and apply the appropriate styling
            if key in ['End']:
                cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wb.save("bid_record.xlsx")
