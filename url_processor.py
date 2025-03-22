from tkinter import messagebox
import openpyxl
import os

class URLProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        self.urls = []
        self.current_idx = -1

    def LoadFile(self):
        """Load the Excel file and initialize sheet."""
        if not os.path.exists(self.file_path):
            # If file doesn't exist, create a blank file
            self.CreateBlankFile()

        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet = self.workbook.active
        except Exception as e:
            print("Error", f"An error occurred while loading the file: {e}")
            return False
        return True

    def CreateBlankFile(self):
        """Create a blank Excel file with an empty sheet."""
        try:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.workbook.save(self.file_path)
        except Exception as e:
            print("Error", f"An error occurred while creating the file: {e}")

    def GetUrls(self):
        """Extract URLs from the first column of the sheet."""
        if not self.sheet:
            messagebox.showerror("Error", "Sheet is not loaded.")
            return []

        self.urls = [(i, self.sheet.cell(row=i, column=1).value) for i in range(1, self.sheet.max_row + 1) if self.sheet.cell(row=i, column=1).value is not None]
        return self.urls

    def SaveFile(self):
        """Save the modified Excel file."""
        try:
            self.workbook.save(self.file_path)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving the file: {e}")

    def First(self):
        """Initialize the iteration over the URLs."""
        self.GetUrls()  # Ensure URLs are loaded
        if self.urls:
            self.current_idx = 0
            return self.urls[self.current_idx]
        else:
            return None

    def Next(self):
        """Move to the next URL in the list."""
        if self.current_idx + 1 < len(self.urls):
            self.current_idx += 1
            return self.urls[self.current_idx]
        else:
            return None
        
    def Prev(self):
        """Move to the previous URL in the list."""
        if self.current_idx - 1 >= 0:
            self.current_idx -= 1
            return self.urls[self.current_idx]
        else:
            return None
    def Cur(self):
        if self.current_idx < len(self.urls) and self.current_idx >= 0:
            return self.urls[self.current_idx]
        else:
            return None
        
    def End(self):
        """Check if the iteration has reached the end."""
        return self.current_idx >= len(self.urls) - 1

    def DelCur(self, row_idx):
        """Delete the current URL from the Excel sheet."""
        if row_idx is not None:
            self.sheet.cell(row=row_idx, column=1).value = None
            print("Result", f"URL at index {row_idx} has been removed from the sheet.")
